import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.axis import ChartLines

# --- Step 1: Pull ALL data with pagination ---
BASE_URL = "https://data.austintexas.gov/resource/9t4d-g238.json"
headers = {"Accept": "application/json"}
limit = 50000
offset = 0
rows = []

while True:
    params = {
        "$limit": limit,
        "$offset": offset,
        "$order": "datetime ASC",
        "$where": "datetime IS NOT NULL"
    }
    resp = requests.get(BASE_URL, headers=headers, params=params)
    resp.raise_for_status()
    batch = resp.json()
    if not batch:
        break
    rows.extend(batch)
    offset += limit

df = pd.DataFrame(rows)

# --- Step 2: Parse date and strip timezone ---
df['date'] = pd.to_datetime(df['datetime'], errors='coerce', utc=True).dt.tz_localize(None)
df = df.dropna(subset=['date'])

# --- Step 3: Filter to dogs only ---
df = df[df['animal_type'].str.lower() == 'dog']

# --- Step 4: Build Month-Year string in MM-YY format ---
df['MonthYear'] = df['date'].dt.strftime('%m-%y')

# --- Step 5: Monthly counts for Adoptions, Transfers, Euthanasia, Return to Owner ---
monthly = df.groupby(['MonthYear','outcome_type']).size().unstack(fill_value=0)
monthly = monthly[['Adoption','Transfer','Euthanasia','Return to Owner']].copy()

# --- Step 6: Sort by actual date order ---
monthly = monthly.reset_index()
monthly['SortDate'] = pd.to_datetime(monthly['MonthYear'], format='%m-%y')
monthly = monthly.sort_values('SortDate').set_index('MonthYear')

# --- Step 7: Compute totals and rates (BEST PRACTICE FIX) ---
monthly['Total'] = monthly.select_dtypes(include='number').sum(axis=1)
monthly['EuthRate'] = monthly['Euthanasia'] / monthly['Total'] * 100

# --- Step 8: Baseline intake before March 2020 ---
baseline_intake = monthly.loc[monthly['SortDate'] < "2020-03-01", 'Total'].mean()

# --- Step 9: Adjusted euthanasia rate (correct formula) ---
monthly['AdjEuthRate'] = (monthly['Euthanasia'] / baseline_intake) * 100

# --- Step 10: Export grouped tables to Excel ---
output_file = "dog_outcomes.xlsx"
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    monthly.to_excel(writer, sheet_name="Dog outcomes raw")

# --- Step 11: Add adjusted euthanasia rate sheet ---
wb = load_workbook(output_file)
ws_adj = wb.create_sheet(title="Adj Euthanasia Rate")

ws_adj.append(["Month", "Total Intake", "Actual Euthanasia Rate %", "Adjusted Euthanasia Rate %"])
for idx, row in monthly.iterrows():
    ws_adj.append([
        idx,
        int(row['Total']),
        round(row['EuthRate'], 2) if pd.notna(row['EuthRate']) else 0,
        round(row['AdjEuthRate'], 2) if pd.notna(row['AdjEuthRate']) else 0
    ])

# --- Step 12: Combined chart (columns + lines) ---
chart = BarChart()
chart.type = "col"
chart.title = "Adjusted Euthanasia Rate vs Intake"
chart.y_axis.title = "Total Intake"
chart.x_axis.title = "Month-Year"

# Intake as columns
data_intake = Reference(ws_adj, min_col=2, min_row=1, max_row=ws_adj.max_row)
cats = Reference(ws_adj, min_col=1, min_row=2, max_row=ws_adj.max_row)
chart.add_data(data_intake, titles_from_data=True)
chart.set_categories(cats)

# Add line chart for rates
line_chart = LineChart()
line_chart.y_axis.title = "Euthanasia Rate (%)"
line_chart.y_axis.scaling.min = 0
line_chart.y_axis.scaling.max = 10
line_chart.y_axis.majorGridlines = ChartLines()

data_rates = Reference(ws_adj, min_col=3, min_row=1, max_col=4, max_row=ws_adj.max_row)
line_chart.add_data(data_rates, titles_from_data=True)
line_chart.set_categories(cats)

# Overlay line chart on bar chart
chart += line_chart

ws_adj.add_chart(chart, "G5")

# --- Step 13: Save workbook ---
wb.save(output_file)

print("Dog data and adjusted euthanasia rate chart exported to dog_outcomes.xlsx")
